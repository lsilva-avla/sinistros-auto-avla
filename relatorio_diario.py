#!/usr/bin/env python3
"""
Relatório Diário de Sinistros - AVLA Crédito
Coleta emails de aviso de sinistro do dia anterior (D-1), enriquece
com tabelas auxiliares (SEGURADOS_CNAE + GRUPO ECONÔMICO), gera Excel,
escreve na base Google Sheets e envia relatório HTML todo dia às 09h BRT.
"""

import imaplib
import email
import smtplib
import re
import os
import sys
import json
import calendar
import urllib.request
import pandas as pd
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email.mime.image import MIMEImage
from email import encoders
from datetime import datetime, timedelta, date
from bs4 import BeautifulSoup
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# =============================================================
# CONFIGURAÇÃO
# =============================================================
EMAIL_CAIXA      = os.environ.get("EMAIL_CAIXA",      "mgignon@avla.com")
APP_PASSWORD     = os.environ.get("APP_PASSWORD",      "")
ASSUNTO_FILTRO   = os.environ.get("ASSUNTO_FILTRO",   "SINISTRO")
REMETENTE_FILTRO = os.environ.get("REMETENTE_FILTRO", "notificaciones-01@avla.com")
_destino_env  = os.environ.get("EMAIL_DESTINO", "")
DESTINATARIOS = [e.strip() for e in _destino_env.split(',') if e.strip()] \
                if _destino_env else ["lsilva@avla.com", "mgignon@avla.com"]
GSHEET_CREDS     = os.environ.get("GSHEET_CREDENTIALS", "")
GSHEET_ID        = os.environ.get("GSHEET_ID", "")

IMAP_SERVER = "imap.gmail.com"
SMTP_SERVER = "smtp.gmail.com"
SMTP_PORT   = 587

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

if not APP_PASSWORD:
    print("ERRO: variável de ambiente APP_PASSWORD não definida.")
    sys.exit(1)
# =============================================================


def periodo_ontem():
    """Retorna (inicio, fim) do dia anterior (D-1)."""
    ontem = date.today() - timedelta(days=1)
    inicio = datetime(ontem.year, ontem.month, ontem.day, 0, 0, 0)
    fim    = datetime(ontem.year, ontem.month, ontem.day, 23, 59, 59)
    return inicio, fim


# ─────────────────────────────────────────────
#  IMAP helpers (iguais ao coletor semanal)
# ─────────────────────────────────────────────

def _select(mail, nome):
    try:
        nome_imap = f'"{nome}"' if any(c in nome for c in '[] ') else nome
        typ, data = mail.select(nome_imap)
        if typ == 'OK':
            count = data[0].decode() if data and data[0] else '?'
            return True, count
        return False, 0
    except Exception:
        return False, 0


def selecionar_pasta_completa(mail):
    try:
        _, pastas = mail.list()
        for item in pastas:
            if item is None:
                continue
            decoded = item.decode('utf-8') if isinstance(item, bytes) else str(item)
            if '\\All' in decoded:
                m = re.search(r'"([^"]+)"\s*$|(\S+)\s*$', decoded)
                if m:
                    nome = (m.group(1) or m.group(2)).strip()
                    ok, count = _select(mail, nome)
                    if ok:
                        print(f"✓ Pasta selecionada (All): {nome} — {count} msgs")
                        return
    except Exception as e:
        print(f"Erro ao detectar All Mail: {e}")

    for nome in ['[Gmail]/All Mail', '[Gmail]/Todos os e-mails', 'INBOX']:
        ok, count = _select(mail, nome)
        if ok:
            print(f"✓ Pasta selecionada (fallback): {nome} — {count} msgs")
            return

    raise RuntimeError("Não foi possível selecionar nenhuma pasta de email.")


def _detectar_lixeira(mail):
    try:
        _, pastas = mail.list()
        for item in pastas:
            if item is None:
                continue
            decoded = item.decode('utf-8') if isinstance(item, bytes) else str(item)
            if '\\Trash' in decoded:
                m = re.search(r'"([^"]+)"\s*$|(\S+)\s*$', decoded)
                if m:
                    return (m.group(1) or m.group(2)).strip()
    except Exception as e:
        print(f"Erro ao detectar Lixeira: {e}")
    return None


# ─────────────────────────────────────────────
#  Parsing de email
# ─────────────────────────────────────────────

def extrair_numero_sinistro(assunto: str) -> str:
    padroes = [
        r'N[°º]?\s*(?:DE\s+)?S[Ii][Nn][Ii][Ee]?[Ss][Tt][Rr][Oo]\s*[:\-]?\s*(\d+)',
        r'S[Ii][Nn][Ii][Ee]?[Ss][Tt][Rr][Oo]\s*[N°nº#]?\s*[:\-]?\s*(\d+)',
        r'\b(\d{6,})\b',
    ]
    for p in padroes:
        m = re.search(p, assunto, re.IGNORECASE)
        if m:
            return m.group(1)
    return ''


def _extrair_num_corpo(html: str) -> str:
    soup = BeautifulSoup(html, 'html.parser')
    texto = soup.get_text(separator=' ')
    padroes = [
        r'N[°º\.]\s*(?:de\s+)?[Ss]ini[es]stro\s*[:\-]?\s*(\d{4,})',
        r'[Ss]ini[es]stro\s*[N°nº#\.]*\s*[:\-]?\s*(\d{4,})',
        r'(?:N[°º]|No\.?|Nro\.?|Nr\.?)\s*[:\-]?\s*(\d{5,})',
        r'\b(\d{8,})\b',
    ]
    for p in padroes:
        m = re.search(p, texto, re.IGNORECASE)
        if m:
            return m.group(1)
    return ''


def parse_corpo_email(html: str) -> dict:
    soup = BeautifulSoup(html, 'html.parser')
    texto = soup.get_text(separator='\n')

    campos_aliases = {
        'N° Sinistro':      ['N° Sinistro', 'Nº Sinistro', 'N° de Sinistro',
                             'Número do Sinistro', 'Número de Sinistro',
                             'N° de Siniestro', 'Nº de Siniestro',
                             'Número de Siniestro', 'No. de Siniestro',
                             'No de Siniestro', 'Siniestro'],
        'Segurado':         ['Segurado'],
        'Filial':           ['Filial'],
        'Apólice':          ['Apólice', 'Apolice', 'Póliza', 'Poliza'],
        'Devedor':          ['Devedor', 'Deudor'],
        'CNPJ do Devedor':  ['CNPJ do devedor', 'CNPJ do Devedor', 'CNPJ',
                             'RUT', 'RUT del Deudor'],
        'Ocorrência':       ['Ocorrência', 'Ocorrencia', 'Siniestro',
                             'Tipo de Siniestro'],
        'Declaração':       ['Declaração', 'Declaracao', 'Fecha de Declaración',
                             'Fecha Declaracion'],
        'Valor Sinistrado': ['Valor sinistrado', 'Valor Sinistrado',
                             'Monto Sinistrado', 'Monto del Siniestro',
                             'Valor do Sinistro'],
    }

    resultado = {}

    for tag in soup.find_all(['b', 'strong']):
        label = tag.get_text(strip=True).rstrip(':')
        for campo, aliases in campos_aliases.items():
            if campo in resultado:
                continue
            for alias in aliases:
                if alias.lower() == label.lower():
                    proximo = tag.next_sibling
                    if proximo:
                        if hasattr(proximo, 'get_text'):
                            valor = proximo.get_text(strip=True).lstrip(':').strip()
                        else:
                            valor = str(proximo).strip().lstrip(':').strip()
                        if valor:
                            resultado[campo] = valor
                    break

    for campo, aliases in campos_aliases.items():
        if campo in resultado:
            continue
        for alias in aliases:
            padrao = re.compile(rf'{re.escape(alias)}\s*:?\s*(.+)', re.IGNORECASE)
            m = padrao.search(texto)
            if m:
                resultado[campo] = m.group(1).strip()
                break

    return resultado


def _buscar_pasta(mail, since, before, vistos):
    try:
        _, ids = mail.search(None, f'(SUBJECT "{ASSUNTO_FILTRO}" SINCE "{since}" BEFORE "{before}")')
    except Exception as e:
        print(f"Erro ao buscar na pasta: {e}")
        return [], vistos

    novos = []
    for msg_id in ids[0].split():
        try:
            _, dados = mail.fetch(msg_id, '(RFC822)')
            msg = email.message_from_bytes(dados[0][1])
        except Exception:
            continue

        mid = msg.get('Message-ID', f'_noid_{msg_id.decode()}_')
        if mid in vistos:
            continue
        vistos.add(mid)

        remetente = str(msg.get('From', ''))
        if REMETENTE_FILTRO and REMETENTE_FILTRO.lower() not in remetente.lower():
            continue

        assunto    = str(msg.get('Subject', ''))
        data_email = email.utils.parsedate_to_datetime(msg['Date'])

        html_body = None
        for part in msg.walk():
            if part.get_content_type() == 'text/html':
                payload   = part.get_payload(decode=True)
                charset   = part.get_content_charset() or 'utf-8'
                html_body = payload.decode(charset, errors='ignore')
                break

        if not html_body:
            continue

        campos      = parse_corpo_email(html_body)
        num_assunto = extrair_numero_sinistro(assunto)
        num_corpo   = campos.get('N° Sinistro', '') or _extrair_num_corpo(html_body)
        num_final   = num_assunto or num_corpo
        if not num_final:
            print(f"  ⚠ N° Sinistro não encontrado | assunto: {assunto[:80]}")
        campos['N° Sinistro'] = num_final
        campos['Data']        = data_email.strftime('%d/%m/%Y')
        novos.append(campos)

    return novos, vistos


def coletar_emails(inicio: datetime, fim: datetime) -> list:
    print(f"Conectando em {EMAIL_CAIXA}...")
    mail = imaplib.IMAP4_SSL(IMAP_SERVER)
    mail.login(EMAIL_CAIXA, APP_PASSWORD)

    since  = inicio.strftime("%d-%b-%Y")
    before = (fim + timedelta(days=1)).strftime("%d-%b-%Y")

    vistos    = set()
    registros = []

    selecionar_pasta_completa(mail)
    novos, vistos = _buscar_pasta(mail, since, before, vistos)
    print(f"All Mail: {len(novos)} sinistros")
    registros.extend(novos)

    nome_lixeira = _detectar_lixeira(mail)
    if nome_lixeira:
        ok, count = _select(mail, nome_lixeira)
        if ok:
            print(f"Lixeira ({nome_lixeira}): {count} msgs — verificando...")
            novos, vistos = _buscar_pasta(mail, since, before, vistos)
            print(f"Lixeira: +{len(novos)} sinistros adicionais")
            registros.extend(novos)

    mail.logout()
    print(f"Total coletado: {len(registros)} sinistros")
    return registros


# ─────────────────────────────────────────────
#  Tabelas auxiliares
# ─────────────────────────────────────────────

def _tc(s) -> str:
    """Title Case seguro: primeira letra maiúscula em cada palavra. Vazio retorna ''."""
    v = str(s).strip() if s is not None else ''
    if not v or v.lower() in ('nan', 'none', ''):
        return ''
    return v.title()


def _normalizar_apolice(v) -> str:
    """
    Normaliza número de apólice para dígitos puros (ex: '1002022000031').
    Aceita: int, float, '1002022000031', '1002022000031.0',
            '1002-022-000031', 'APL-1002022000031', '1002 022 000031'.
    """
    s = str(v).strip() if v is not None else ''
    if not s or s.lower() in ('nan', 'none', ''):
        return ''
    # Tenta int(float) primeiro (resolve o '.0' de células numéricas)
    try:
        return str(int(float(s)))
    except (ValueError, TypeError):
        pass
    # Fallback: extrai só os dígitos (resolve traços, prefixos, espaços internos)
    return re.sub(r'\D', '', s)


def _normalizar_cnpj(v) -> str:
    """
    Normaliza CNPJ para 14 dígitos puros com zero à esquerda preservado.
    Aceita: '03.439.316/0001-72', '03439316000172', '3439316000172' (faltando 0).
    """
    digits = re.sub(r'\D', '', str(v or ''))
    if not digits:
        return ''
    # CNPJ tem 14 dígitos; CPF tem 11 — garante zero-padding correto
    if len(digits) == 13:          # zero à esquerda foi perdido
        digits = digits.zfill(14)
    elif len(digits) == 10:        # CPF sem zero à esquerda
        digits = digits.zfill(11)
    return digits


def _col_letter(n: int) -> str:
    """Converte índice de coluna 1-based em letra(s) Excel. 1→A, 26→Z, 27→AA."""
    s = ''
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def _find_col(df, *keywords):
    """
    Localiza coluna pelo nome com 3 níveis de precisão (case-insensitive):
      1. Correspondência exata
      2. Nome da coluna COMEÇA COM o keyword (evita 'Id Grupo Econ.' bater em 'grupo econ')
      3. Contém o keyword (fallback amplo)
    Aceita múltiplos keywords alternativos — retorna a primeira coluna que bater.
    """
    cols_lower = {str(c).lower(): c for c in df.columns}
    for kw in keywords:
        kw_l = kw.lower()
        # Nível 1 — exato
        if kw_l in cols_lower:
            return cols_lower[kw_l]
    for kw in keywords:
        kw_l = kw.lower()
        # Nível 2 — começa com keyword
        for cl, c in cols_lower.items():
            if cl.startswith(kw_l):
                return c
    for kw in keywords:
        kw_l = kw.lower()
        # Nível 3 — contém keyword (fallback)
        for cl, c in cols_lower.items():
            if kw_l in cl:
                return c
    return None


def carregar_tabelas_auxiliares():
    """Carrega lookups de SEGURADOS_CNAE e GRUPO_ECONOMICO."""
    path_cnae  = os.path.join(BASE_DIR, 'dados', 'SEGURADOS_CNAE.xlsx')
    path_grupo = os.path.join(BASE_DIR, 'dados', 'GRUPO_ECONOMICO.xlsx')

    lookup_cnae  = {}
    lookup_grupo = {}

    # — SEGURADOS_CNAE —
    if os.path.exists(path_cnae):
        try:
            df = pd.read_excel(path_cnae)
            # Localiza colunas por nome — robusto contra renomeações
            col_pol  = _find_col(df, 'poliza', 'apolice', 'apólice') or df.columns[0]
            col_vi   = _find_col(df, 'vigencia inicio', 'vigência inicio', 'inicio') or 'Vigencia Inicio'
            col_vf   = _find_col(df, 'vigencia fim',    'vigência fim',    'fim')    or 'Vigencia Fim'
            col_set  = _find_col(df, 'setor')    or 'SETOR'
            col_sub  = _find_col(df, 'subsetor') or 'SUBSETOR'
            print(f"  > CNAE cols: pol={col_pol}, vi={col_vi}, vf={col_vf}, setor={col_set}, sub={col_sub}")
            df[col_pol] = df[col_pol].apply(_normalizar_apolice)
            for _, row in df.iterrows():
                chave = row[col_pol]
                if not chave:
                    continue
                vi = row.get(col_vi, '')
                vf = row.get(col_vf, '')
                lookup_cnae[chave] = {
                    'Vigencia Inicio': vi.strftime('%d/%m/%Y') if hasattr(vi, 'strftime') else str(vi or ''),
                    'Vigencia Fim':    vf.strftime('%d/%m/%Y') if hasattr(vf, 'strftime') else str(vf or ''),
                    'SETOR':    _tc(row.get(col_set, '')),
                    'SUBSETOR': _tc(row.get(col_sub, '')),
                }
            print(f"SEGURADOS_CNAE: {len(lookup_cnae)} apólices carregadas.")
        except Exception as e:
            print(f"Erro ao carregar SEGURADOS_CNAE: {e}")
    else:
        print("SEGURADOS_CNAE.xlsx nao encontrado — enriquecimento CNAE pulado.")

    # — GRUPO ECONÔMICO — (gravado no Sheets para o dashboard; fora do email/Excel)
    if os.path.exists(path_grupo):
        try:
            df = pd.read_excel(path_grupo)
            col_id    = _find_col(df, 'identificador participante', 'identificador') or df.columns[2]
            col_grupo = _find_col(df, 'grupo econ', 'grupo economico') or df.columns[4]
            print(f"  > Grupo cols: id={col_id}, grupo={col_grupo}")
            for _, row in df.iterrows():
                digits = _normalizar_cnpj(row[col_id])
                grupo  = _tc(row[col_grupo])
                if digits and grupo and digits not in lookup_grupo:
                    lookup_grupo[digits] = grupo
            print(f"GRUPO ECONOMICO: {len(lookup_grupo)} participantes carregados.")
        except Exception as e:
            print(f"Erro ao carregar GRUPO_ECONOMICO: {e}")
    else:
        print("GRUPO_ECONOMICO.xlsx nao encontrado.")

    return lookup_cnae, lookup_grupo


def enriquecer(registros: list, lookup_cnae: dict, lookup_grupo: dict) -> list:
    """
    Enriquece cada registro com:
    - Title Case nos campos de texto livre
    - CNAE: Vigencia Inicio/Fim, SETOR, SUBSETOR (via apólice)
    - Grupo Econômico (via CNPJ — apenas no Sheets/dashboard, fora do email)
    - Data_ISO (YYYY-MM-DD) e Mes_Ano (YYYY-MM) para filtros no Looker
    - Valor BRL (float), FX PTAX e Valor USD (conversão pelo último dia útil do mês)
    """
    sem_cnae = []

    for r in registros:
        # Title Case em todos os campos de texto livre vindos do email
        for campo in ('Segurado', 'Filial', 'Devedor', 'Ocorrência', 'Declaração'):
            if r.get(campo):
                r[campo] = _tc(r[campo])

        # SEGURADOS_CNAE: apólice normalizada para dígitos puros
        apolice   = _normalizar_apolice(r.get('Apólice', ''))
        cnae_info = lookup_cnae.get(apolice, {})
        r['Vigencia Inicio'] = cnae_info.get('Vigencia Inicio', '')
        r['Vigencia Fim']    = cnae_info.get('Vigencia Fim', '')
        r['SETOR']           = cnae_info.get('SETOR', '')
        r['SUBSETOR']        = cnae_info.get('SUBSETOR', '')
        if not cnae_info:
            sem_cnae.append(apolice or r.get('Apólice', '?'))

        # Grupo Econômico via CNPJ
        cnpj_digits          = _normalizar_cnpj(r.get('CNPJ do Devedor', ''))
        r['Grupo Econômico'] = lookup_grupo.get(cnpj_digits, '')

        # Data no formato ISO e agrupamento Mês/Ano
        d = _parse_date(r.get('Data', ''))
        r['Data_ISO'] = d.strftime('%Y-%m-%d') if d.year >= 2020 else ''
        r['Mes_Ano']  = d.strftime('%Y-%m')    if d.year >= 2020 else ''

        # Detecta moeda e extrai valor original; converte para BRL se necessário
        moeda, valor_original = _detect_moeda(r.get('Valor Sinistrado', ''))
        r['Moeda']          = moeda
        r['Valor Original'] = round(valor_original, 2)

        if d.year >= 2020 and valor_original:
            if moeda == 'BRL':
                ptax_usd       = _get_ptax(d.year, d.month, 'USD')
                r['Valor BRL'] = round(valor_original, 2)
                r['FX PTAX']   = ptax_usd
                r['Valor USD'] = round(valor_original / ptax_usd, 2) if ptax_usd else 0.0
            elif moeda == 'USD':
                ptax_usd       = _get_ptax(d.year, d.month, 'USD')
                r['Valor BRL'] = round(valor_original * ptax_usd, 2) if ptax_usd else 0.0
                r['FX PTAX']   = ptax_usd
                r['Valor USD'] = round(valor_original, 2)
            elif moeda == 'EUR':
                ptax_eur       = _get_ptax(d.year, d.month, 'EUR')
                ptax_usd       = _get_ptax(d.year, d.month, 'USD')
                valor_brl      = round(valor_original * ptax_eur, 2) if ptax_eur else 0.0
                r['Valor BRL'] = valor_brl
                r['FX PTAX']   = ptax_usd  # referência USD para Valor USD
                r['Valor USD'] = round(valor_brl / ptax_usd, 2) if ptax_usd else 0.0
            else:
                r['Valor BRL'] = 0.0
                r['FX PTAX']   = 0.0
                r['Valor USD'] = 0.0
        else:
            r['Valor BRL'] = 0.0
            r['FX PTAX']   = 0.0
            r['Valor USD'] = 0.0

        # Campos de rastreabilidade
        r.setdefault('Origem', 'Email')
        r.setdefault('Observações', '')

    total    = len(registros)
    com_cnae = total - len(sem_cnae)
    print(f"Enriquecimento: {com_cnae}/{total} com CNAE")
    if sem_cnae:
        print(f"  [!] Apolices sem match na tabela: {sorted(set(sem_cnae))}")

    return registros


def _parse_date(s: str) -> date:
    """Converte string dd/mm/yyyy para date; retorna date mínima em caso de erro."""
    try:
        return datetime.strptime(s, '%d/%m/%Y').date()
    except (ValueError, TypeError):
        return date(2000, 1, 1)


# ─────────────────────────────────────────────
#  Cotação FX — BCB PTAX (gratuita, sem chave)
# ─────────────────────────────────────────────

_ptax_cache: dict = {}

def _get_ptax(ano: int, mes: int, moeda: str = 'USD') -> float:
    """
    Cotação BRL por 1 unidade da moeda (PTAX venda) do último dia útil do mês via BCB.
    Suporta: USD (padrão), EUR e demais moedas aceitas pelo BCB.
    Para o mês corrente, usa o último dia disponível (até ontem).
    Cache por (moeda, ano, mes) para evitar chamadas repetidas.
    """
    chave = (moeda, ano, mes)
    if chave in _ptax_cache:
        return _ptax_cache[chave]

    hoje   = date.today()
    ultimo = calendar.monthrange(ano, mes)[1]
    if ano == hoje.year and mes == hoje.month:
        ultimo = min(ultimo, (hoje - timedelta(days=1)).day)

    for delta in range(0, 10):
        d = date(ano, mes, ultimo) - timedelta(days=delta)
        if d.month != mes:
            break
        data_str = d.strftime('%m-%d-%Y')
        if moeda == 'USD':
            url = (
                'https://olinda.bcb.gov.br/olinda/servico/PTAX/versao/v1/odata/'
                f'CotacaoDolarDia(dataCotacao=@d)?@d=%27{data_str}%27'
                '&$format=json&$select=cotacaoVenda'
            )
        else:
            url = (
                'https://olinda.bcb.gov.br/olinda/servico/PTAX/versao/v1/odata/'
                f'CotacaoMoedaDia(moeda=@m,dataCotacao=@d)?@m=%27{moeda}%27&@d=%27{data_str}%27'
                '&$format=json&$select=cotacaoVenda'
            )
        try:
            req = urllib.request.Request(url, headers={'Accept': 'application/json'})
            with urllib.request.urlopen(req, timeout=8) as resp:
                valores = json.loads(resp.read()).get('value', [])
                if valores:
                    taxa = float(valores[-1]['cotacaoVenda'])
                    _ptax_cache[chave] = taxa
                    print(f"  PTAX {moeda} {ano}/{mes:02d}: R$ {taxa:.4f} (ref. {d})")
                    return taxa
        except Exception:
            pass

    print(f"  [!] PTAX {moeda} nao encontrada para {ano}/{mes:02d}")
    _ptax_cache[chave] = 0.0
    return 0.0


def _detect_moeda(s: str):
    """
    Detecta a moeda e extrai o valor numérico de uma string de valor monetário.
    Suporta formato brasileiro: ponto = separador de milhar, vírgula = decimal.

    Exemplos:
      "R$ 192.434,86"   -> ('BRL', 192434.86)
      "US$ 37.668,24"   -> ('USD', 37668.24)
      "USS 11.210,85"   -> ('USD', 11210.85)
      "2.321.118,00 €"  -> ('EUR', 2321118.0)
      "USD 1.234,56"    -> ('USD', 1234.56)

    Retorna (moeda: str, valor: float)
    """
    s = str(s).strip() if s else ''
    if not s or s.lower() in ('nan', 'none', ''):
        return 'BRL', 0.0

    s_upper = s.upper()

    # Detecta moeda pela presença de símbolos / siglas
    if any(tok in s_upper for tok in ('US$', 'USS', 'USD', 'U$S')):
        moeda = 'USD'
    elif '€' in s or 'EUR' in s_upper:
        moeda = 'EUR'
    else:
        moeda = 'BRL'   # R$, RS, BRL, ou sem prefixo

    # Extrai apenas dígitos, ponto e vírgula
    nums = re.sub(r'[^\d.,]', '', s)
    if not nums:
        return moeda, 0.0

    # Formato BR: vírgula = decimal, ponto = milhar
    if ',' in nums:
        nums = nums.replace('.', '').replace(',', '.')
    elif nums.count('.') > 1:
        # Múltiplos pontos sem vírgula → todos são separadores de milhar
        nums = nums.replace('.', '')
    elif '.' in nums:
        # Um único ponto sem vírgula — heurística: 3 dígitos após ponto = milhar
        after_dot = nums.split('.')[1]
        if len(after_dot) == 3:
            nums = nums.replace('.', '')
        # else: trata o ponto como decimal (ex: "1.5" = 1.5)

    try:
        return moeda, float(nums)
    except (ValueError, TypeError):
        return moeda, 0.0


# ─────────────────────────────────────────────
#  Google Sheets
# ─────────────────────────────────────────────

COLUNAS_SHEETS = [
    'ID', 'N° Sinistro', 'Data', 'Data_ISO', 'Mes_Ano',
    'Segurado', 'Filial', 'Apólice',
    'Devedor', 'CNPJ do Devedor', 'Ocorrência', 'Declaração',
    'Valor Sinistrado', 'Moeda', 'Valor Original', 'Valor BRL', 'FX PTAX', 'Valor USD',
    'Vigencia Inicio', 'Vigencia Fim',
    'SETOR', 'SUBSETOR', 'Grupo Econômico',
    'Observações', 'Origem',           # 'Observações' = notas manuais; 'Origem' = Email|Manual
]

# Colunas que devem ser gravadas como número (float) no Sheets
_COLUNAS_NUMERICAS = {'Valor Original', 'Valor BRL', 'FX PTAX', 'Valor USD'}


# ─────────────────────────────────────────────
#  Casos manuais (segunda aba do Sheets)
# ─────────────────────────────────────────────

def _mapear_header_manual(header: list) -> dict:
    """
    Mapeia as colunas da aba manual para os campos internos de COLUNAS_SHEETS.
    Retorna {indice_coluna: nome_campo_interno}.
    Usa matching flexível: exato > startswith > contains.
    """
    _alias = {
        'N° Sinistro':      ['n° sinistro', 'nº sinistro', 'n sinistro', 'sinistro',
                             'no sinistro', 'numero sinistro', 'num sinistro', 'n°sinistro'],
        'Data':             ['data', 'date', 'fecha'],
        'Segurado':         ['segurado'],
        'Filial':           ['filial'],
        'Apólice':          ['apolice', 'apólice', 'poliza', 'póliza', 'policia'],
        'Devedor':          ['devedor', 'deudor', 'tomador'],
        'CNPJ do Devedor':  ['cnpj', 'cpf/cnpj', 'cpf', 'rut', 'documento'],
        'Ocorrência':       ['ocorrência', 'ocorrencia', 'tipo', 'tipo sinistro',
                             'tipo de sinistro'],
        'Declaração':       ['declaração', 'declaracao', 'fecha declaracion'],
        'Valor Sinistrado': ['valor sinistrado', 'valor', 'monto', 'monto sinistrado',
                             'valor do sinistro'],
        'Observações':      ['observações', 'observacao', 'observacoes', 'obs',
                             'notas', 'nota', 'descricao', 'descrição',
                             'dicionario', 'dicionário', 'comentario', 'comentários'],
    }
    mapping = {}
    campos_mapeados = set()
    for ci, col in enumerate(header):
        col_l = col.strip().lower()
        if not col_l:
            continue
        for campo, aliases in _alias.items():
            if campo in campos_mapeados:
                continue
            for a in aliases:
                if col_l == a or col_l.startswith(a) or a in col_l:
                    mapping[ci]       = campo
                    campos_mapeados.add(campo)
                    break
    return mapping


def ler_casos_manuais(sh, lookup_cnae: dict, lookup_grupo: dict) -> list:
    """
    Lê a segunda aba do Sheets (qualquer aba que não seja 'Base').
    - Filtra linhas marcadas com X se houver coluna de marcação
    - Enriquece sem sobrescrever campos já preenchidos por Maity
    - Retorna registros com Origem='Manual'
    """
    try:
        worksheets  = sh.worksheets()
        aba_manual  = next((ws for ws in worksheets
                            if ws.title.strip().lower() != 'base'), None)
        if aba_manual is None:
            print("  [!] Aba manual nao encontrada (so existe 'Base').")
            return []

        print(f"  > Lendo casos manuais de '{aba_manual.title}'...")
        rows = aba_manual.get_all_values()
        if not rows or len(rows) < 2:
            print("  Aba manual vazia.")
            return []

        header  = rows[0]
        mapping = _mapear_header_manual(header)
        print(f"  > Colunas mapeadas: {list(mapping.values())}")

        # Importa TODAS as linhas não-vazias — toda a 2ª aba é manual
        registros = []
        for row in rows[1:]:
            if not any(c.strip() for c in row):
                continue   # pula linha completamente vazia

            r = {}
            for ci, campo in mapping.items():
                val = row[ci].strip() if len(row) > ci else ''
                if val and val.lower() not in ('nan', 'none'):
                    r[campo] = val

            # Precisa ter pelo menos N° Sinistro ou Devedor
            if not r.get('N° Sinistro') and not r.get('Devedor'):
                continue

            r['Origem']     = 'Manual'
            r.setdefault('Observações', '')
            registros.append(r)

        print(f"  Casos manuais lidos: {len(registros)} de {len(rows)-1} linhas.")

        # ── Enriquecimento sem sobrescrever campos manuais ──────
        for r in registros:
            for campo in ('Segurado', 'Filial', 'Devedor', 'Ocorrência', 'Declaração'):
                if r.get(campo):
                    r[campo] = _tc(r[campo])

            apolice   = _normalizar_apolice(r.get('Apólice', ''))
            cnae_info = lookup_cnae.get(apolice, {})
            for chave in ('Vigencia Inicio', 'Vigencia Fim', 'SETOR', 'SUBSETOR'):
                if not r.get(chave):
                    r[chave] = cnae_info.get(chave, '')

            if not r.get('Grupo Econômico'):
                cnpj_d = _normalizar_cnpj(r.get('CNPJ do Devedor', ''))
                r['Grupo Econômico'] = lookup_grupo.get(cnpj_d, '')

            d = _parse_date(r.get('Data', ''))
            r.setdefault('Data_ISO', d.strftime('%Y-%m-%d') if d.year >= 2020 else '')
            r.setdefault('Mes_Ano',  d.strftime('%Y-%m')    if d.year >= 2020 else '')

            if not r.get('Valor BRL'):
                moeda, v_orig = _detect_moeda(r.get('Valor Sinistrado', ''))
                r.setdefault('Moeda',          moeda)
                r.setdefault('Valor Original', round(v_orig, 2))
                if d.year >= 2020 and v_orig:
                    if moeda == 'BRL':
                        pu = _get_ptax(d.year, d.month, 'USD')
                        r['Valor BRL'] = round(v_orig, 2)
                        r['FX PTAX']   = pu
                        r['Valor USD'] = round(v_orig / pu, 2) if pu else 0.0
                    elif moeda == 'USD':
                        pu = _get_ptax(d.year, d.month, 'USD')
                        r['Valor BRL'] = round(v_orig * pu, 2) if pu else 0.0
                        r['FX PTAX']   = pu
                        r['Valor USD'] = round(v_orig, 2)
                    elif moeda == 'EUR':
                        pe = _get_ptax(d.year, d.month, 'EUR')
                        pu = _get_ptax(d.year, d.month, 'USD')
                        vb = round(v_orig * pe, 2) if pe else 0.0
                        r['Valor BRL'] = vb
                        r['FX PTAX']   = pu
                        r['Valor USD'] = round(vb / pu, 2) if pu else 0.0

        return registros

    except Exception as e:
        print(f"  [!] Erro ao ler casos manuais: {e}")
        return []


# ─────────────────────────────────────────────
#  Google Sheets — Base
# ─────────────────────────────────────────────

# Colunas enriquecidas/calculadas → cabeçalho verde no Sheets (espelha o Excel)
_COLS_SHEETS_VERDE = {
    'Data_ISO', 'Mes_Ano',
    'Moeda', 'Valor Original', 'Valor BRL', 'FX PTAX', 'Valor USD',
    'Vigencia Inicio', 'Vigencia Fim',
    'SETOR', 'SUBSETOR', 'Grupo Econômico',
    'Origem',
}


def escrever_sheets(registros: list, lookup_cnae: dict = None, lookup_grupo: dict = None):
    """
    Reconstrói a aba 'Base' do zero a cada execução com layout igual ao Excel:
      Linha 1 : título mesclado (azul escuro, branco) — idêntico ao Excel
      Linha 2 : cabeçalhos de coluna (azul / verde por categoria)
      Linha 3+: dados, ordenados por data (mais antiga → mais recente)
    Casos manuais da 2ª aba ficam em laranja.
    Toda a formatação é aplicada em um único batchUpdate.
    """
    if not GSHEET_CREDS or not GSHEET_ID:
        print("Sheets: credenciais nao configuradas — pulando.")
        return

    lookup_cnae  = lookup_cnae  or {}
    lookup_grupo = lookup_grupo or {}

    try:
        import gspread
        from google.oauth2.service_account import Credentials

        creds = Credentials.from_service_account_info(
            json.loads(GSHEET_CREDS),
            scopes=[
                'https://spreadsheets.google.com/feeds',
                'https://www.googleapis.com/auth/drive',
            ]
        )
        gc = gspread.authorize(creds)
        sh = gc.open_by_key(GSHEET_ID)

        try:
            aba = sh.worksheet('Base')
        except gspread.exceptions.WorksheetNotFound:
            aba = sh.add_worksheet('Base', rows=50000, cols=len(COLUNAS_SHEETS))

        # ── Casos manuais da 2ª aba ───────────────────────────
        manuais = ler_casos_manuais(sh, lookup_cnae, lookup_grupo)

        # ── Merge: email tem prioridade no dedup ──────────────
        num_emails = {str(r.get('N° Sinistro', '')) for r in registros}
        extras     = [m for m in manuais if str(m.get('N° Sinistro', '')) not in num_emails]
        todos      = list(registros) + extras

        # ── Ordena por Data_ISO (mais antiga → mais recente) ──
        todos.sort(key=lambda r: r.get('Data_ISO', '') or 'zzzz')

        # ── Monta linhas de dados ─────────────────────────────
        # Linha 1 = título | Linha 2 = cabeçalhos | Linha 3+ = dados
        rows_data      = []
        linhas_manuais = []

        for seq_id, r in enumerate(todos, start=1):
            row = [seq_id]
            for c in COLUNAS_SHEETS[1:]:
                val = r.get(c, '') or ''
                if c in _COLUNAS_NUMERICAS:
                    try:
                        row.append(float(val))
                    except (ValueError, TypeError):
                        row.append(0.0)
                else:
                    row.append(str(val))
            rows_data.append(row)
            if r.get('Origem') == 'Manual':
                linhas_manuais.append(seq_id + 2)  # +2: linha 1=título, linha 2=cabeçalho

        # ── Reconstrói Base (limpa e reescreve tudo de uma vez) ──
        aba.clear()
        titulo    = f'Relatório de Sinistros — AVLA Crédito {date.today().year}'
        linha_tit = [titulo] + [''] * (len(COLUNAS_SHEETS) - 1)
        aba.update('A1', [linha_tit, COLUNAS_SHEETS] + rows_data,
                   value_input_option='USER_ENTERED')

        n_email  = len(todos) - len(extras)
        n_manual = len(extras)
        print(f"Sheets: Base reconstruida — {len(todos)} registros "
              f"({n_email} email + {n_manual} manuais), ordenados por data.")

        # ── Formatação em batch (1 chamada à API) ─────────────
        nc  = len(COLUNAS_SHEETS)
        sid = aba.id

        def _rgb(r, g, b):
            return {'red': round(r/255, 4), 'green': round(g/255, 4), 'blue': round(b/255, 4)}

        C_AZUL    = _rgb(0,   48,  135)   # #003087
        C_VERDE   = _rgb(29,  111,  66)   # #1D6F42
        C_BRANCO  = _rgb(255, 255, 255)
        C_LARANJA = _rgb(255, 153,   0)   # laranja

        def _rng(r0, r1, c0=0, c1=None):
            return {'sheetId': sid, 'startRowIndex': r0, 'endRowIndex': r1,
                    'startColumnIndex': c0, 'endColumnIndex': c1 or nc}

        requests = []

        # 1. Desfaz mesclagem anterior e aplica nova mescla no título
        requests += [
            {'unmergeCells': {'range': _rng(0, 1)}},
            {'mergeCells':   {'range': _rng(0, 1), 'mergeType': 'MERGE_ALL'}},
        ]

        # 2. Formata linha 1 — título
        requests.append({
            'repeatCell': {
                'range': _rng(0, 1),
                'cell': {'userEnteredFormat': {
                    'backgroundColor':    C_AZUL,
                    'horizontalAlignment': 'CENTER',
                    'verticalAlignment':   'MIDDLE',
                    'textFormat': {'foregroundColor': C_BRANCO,
                                   'bold': True, 'fontSize': 13,
                                   'fontFamily': 'Arial'},
                }},
                'fields': ('userEnteredFormat(backgroundColor,horizontalAlignment,'
                           'verticalAlignment,textFormat)'),
            }
        })

        # 3. Formata linha 2 — cabeçalhos (azul ou verde por coluna)
        for ci, col in enumerate(COLUNAS_SHEETS):
            bg = C_VERDE if col in _COLS_SHEETS_VERDE else C_AZUL
            requests.append({
                'repeatCell': {
                    'range': _rng(1, 2, ci, ci + 1),
                    'cell': {'userEnteredFormat': {
                        'backgroundColor':    bg,
                        'horizontalAlignment': 'CENTER',
                        'verticalAlignment':   'MIDDLE',
                        'wrapStrategy':        'WRAP',
                        'textFormat': {'foregroundColor': C_BRANCO,
                                       'bold': True, 'fontSize': 10,
                                       'fontFamily': 'Arial'},
                    }},
                    'fields': ('userEnteredFormat(backgroundColor,horizontalAlignment,'
                               'verticalAlignment,wrapStrategy,textFormat)'),
                }
            })

        # 4. Congela as 2 primeiras linhas (título + cabeçalho)
        requests.append({
            'updateSheetProperties': {
                'properties': {'sheetId': sid,
                               'gridProperties': {'frozenRowCount': 2}},
                'fields': 'gridProperties.frozenRowCount',
            }
        })

        # 5. Altura das linhas 1 e 2
        requests.append({
            'updateDimensionProperties': {
                'range':      {'sheetId': sid, 'dimension': 'ROWS',
                               'startIndex': 0, 'endIndex': 2},
                'properties': {'pixelSize': 28},
                'fields':     'pixelSize',
            }
        })

        # 6. Formato moeda R$ na coluna Valor BRL (dados: a partir da linha 3 → índice 2)
        col_brl = COLUNAS_SHEETS.index('Valor BRL')
        requests.append({
            'repeatCell': {
                'range': _rng(2, 50000, col_brl, col_brl + 1),
                'cell': {'userEnteredFormat': {
                    'numberFormat': {'type': 'CURRENCY', 'pattern': '"R$ "#,##0.00'}
                }},
                'fields': 'userEnteredFormat.numberFormat',
            }
        })

        # 7. Laranja nos casos manuais
        for ln in linhas_manuais:
            requests.append({
                'repeatCell': {
                    'range': _rng(ln - 1, ln),
                    'cell': {'userEnteredFormat': {'backgroundColor': C_LARANJA}},
                    'fields': 'userEnteredFormat.backgroundColor',
                }
            })

        sh.batch_update({'requests': requests})
        print(f"  > Formatação aplicada: título, cabeçalhos, "
              f"{len(linhas_manuais)} linha(s) laranja.")

    except Exception as e:
        print(f"Erro ao escrever no Sheets (nao fatal): {e}")


# ─────────────────────────────────────────────
#  Excel
# ─────────────────────────────────────────────

def gerar_excel(registros: list, ontem: date, fallback: bool = False) -> BytesIO:
    """
    Gera Excel com TODOS os registros do ano (23 colunas, A–W).
    - Linha 1 : título mesclado A1:W1 — azul escuro, texto branco
    - Linha 2 : cabeçalhos de coluna — azul (dados) / verde (enriquecidos), texto branco
    - Linha 3+: dados; destaca ontem (normal) ou últimos 7 dias (fallback)
    """
    colunas = list(COLUNAS_SHEETS)   # 23 colunas (A–W)

    # ── Monta DataFrame ──────────────────────────────────────
    df = pd.DataFrame(registros)
    for col in colunas:
        if col not in df.columns:
            df[col] = ''
    df = df.reindex(columns=[c for c in colunas if c != 'ID'])

    # ── Paleta / estilos ─────────────────────────────────────
    AZUL    = '003087'
    VERDE   = '1D6F42'
    BRANCO  = 'FFFFFF'
    DEST    = 'DCF0FF'

    fill_azul   = PatternFill(fill_type='solid', fgColor=AZUL)
    fill_verde  = PatternFill(fill_type='solid', fgColor=VERDE)
    fill_dest   = PatternFill(fill_type='solid', fgColor=DEST)
    font_titulo = Font(bold=True, color=BRANCO, size=13, name='Arial')
    font_header = Font(bold=True, color=BRANCO, size=10, name='Arial')
    font_body   = Font(name='Arial', size=10)
    al_center   = Alignment(horizontal='center', vertical='center')
    al_wrap     = Alignment(horizontal='center', vertical='center', wrap_text=True)
    al_body     = Alignment(vertical='center')

    # Colunas enriquecidas/calculadas → cabeçalho verde
    cols_verde = {
        'Data_ISO', 'Mes_Ano',
        'Moeda', 'Valor Original', 'Valor BRL', 'FX PTAX', 'Valor USD',
        'Vigencia Inicio', 'Vigencia Fim',
        'SETOR', 'SUBSETOR', 'Grupo Econômico',
        'Origem',   # derivado do processamento
    }
    # Colunas numéricas → gravar como float + formato numérico
    fmt_map = {
        'Valor Original': '#,##0.00',
        'Valor BRL':      '#,##0.00',
        'FX PTAX':        '#,##0.0000',
        'Valor USD':      '#,##0.00',
    }

    # ── Workbook ─────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sinistros 2026'

    num_cols   = len(colunas)
    ultima_col = _col_letter(num_cols)   # 23 → 'W'

    # Linha 1 — título mesclado ───────────────────────────────
    ws.merge_cells(f'A1:{ultima_col}1')
    tc = ws['A1']
    tc.value     = f'Relatório de Sinistros — AVLA Crédito {ontem.year}'
    tc.font      = font_titulo
    tc.fill      = fill_azul
    tc.alignment = al_center
    ws.row_dimensions[1].height = 30

    # Linha 2 — cabeçalhos ────────────────────────────────────
    for ci, col_name in enumerate(colunas, start=1):
        cell = ws.cell(row=2, column=ci, value=col_name)
        cell.font      = font_header
        cell.fill      = fill_verde if col_name in cols_verde else fill_azul
        cell.alignment = al_wrap
    ws.row_dimensions[2].height = 24

    # Lógica de destaque ──────────────────────────────────────
    ontem_str = ontem.strftime('%d/%m/%Y')
    corte_7d  = date.today() - timedelta(days=7)

    # Campos de texto → Title Case no Excel (defesa além do enriquecer)
    cols_texto = {
        'Segurado', 'Filial', 'Devedor', 'Ocorrência', 'Declaração',
        'SETOR', 'SUBSETOR', 'Grupo Econômico', 'Observações',
    }

    # Linha 3+ — dados ────────────────────────────────────────
    for seq_id, (_, row) in enumerate(df.iterrows(), start=1):
        xl_row = seq_id + 2   # linha Excel = seq_id + 2 (linhas 1 e 2 são cabeçalhos)

        data_str = str(row.get('Data', '') or '')
        try:
            data_row = datetime.strptime(data_str, '%d/%m/%Y').date()
            destacar = (data_row >= corte_7d) if fallback else (data_str == ontem_str)
        except (ValueError, TypeError):
            destacar = False

        fill_linha = fill_dest if destacar else None

        for ci, col_name in enumerate(colunas, start=1):
            if col_name == 'ID':
                val = seq_id
            else:
                raw = row.get(col_name, '')
                if col_name in fmt_map:
                    try:
                        val = float(raw) if raw not in ('', None) else 0.0
                    except (ValueError, TypeError):
                        val = 0.0
                else:
                    val = '' if (raw is None or str(raw).lower() in ('nan', 'none', '')) else str(raw)
                    if val and col_name in cols_texto:
                        val = _tc(val)

            cell = ws.cell(row=xl_row, column=ci, value=val)
            cell.font      = font_body
            cell.alignment = al_body
            if col_name in fmt_map:
                cell.number_format = fmt_map[col_name]
            if fill_linha:
                cell.fill = fill_linha

    # Auto-width ──────────────────────────────────────────────
    sample = df.head(50)
    for ci, col_name in enumerate(colunas, start=1):
        letra = _col_letter(ci)
        if col_name in sample.columns and not sample.empty:
            amostra = sample[col_name].fillna('').astype(str).tolist()
        else:
            amostra = []
        max_len = max(len(col_name), max((len(v) for v in amostra), default=0))
        ws.column_dimensions[letra].width = min(max_len + 4, 40)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─────────────────────────────────────────────
#  Email HTML
# ─────────────────────────────────────────────

def _logo_path():
    return os.path.join(BASE_DIR, 'avla_logo.png')


def parse_valor(s: str) -> float:
    try:
        s = re.sub(r'[^\d,]', '', str(s))
        return float(s.replace(',', '.')) if s else 0.0
    except Exception:
        return 0.0


def formatar_brl(v: float) -> str:
    if v >= 1_000_000:
        return 'R$ ' + '{:.1f}'.format(v / 1_000_000).replace('.', ',') + 'M'
    if v >= 1_000:
        return 'R$ ' + '{:.0f}'.format(v / 1_000) + 'K'
    return 'R$ ' + '{:.2f}'.format(v).replace('.', ',')


def gerar_html_email(qtd, registros, label_periodo, nome_arquivo,
                     tem_logo=True, fallback=False, casos_ontem=0, total_ano=0):
    img_h = ('<img src="cid:avla_logo" alt="AVLA" style="height:52px;display:block;margin:0 auto;">'
             if tem_logo else '<span style="color:white;font-size:26px;font-weight:bold;">AVLA</span>')
    img_f = ('<img src="cid:avla_logo" alt="AVLA" style="height:36px;display:block;margin:0 auto;">'
             if tem_logo else '<span style="color:white;font-size:18px;font-weight:bold;">AVLA</span>')

    total_v = sum(float(r.get('Valor BRL') or 0) for r in registros)
    maior_v = max((float(r.get('Valor BRL') or 0) for r in registros), default=0.0)
    maior_r = next((r for r in registros if float(r.get('Valor BRL') or 0) == maior_v), {})
    maior_nome = maior_r.get('Devedor', maior_r.get('Segurado', '—'))
    if len(maior_nome) > 22:
        maior_nome = maior_nome[:22] + '…'

    def card(cor, label, val, sub, width='33%', destaque_cor=None):
        fs = '22' if width == '33%' and label in ('Total de Casos', 'Ontem') else '18'
        val_style = f'color:{destaque_cor};' if destaque_cor else 'color:#1a1a1a;'
        return (
            f'<td width="{width}" style="padding:0 5px;">'
            f'<div style="border:1px solid #e0e0e0;border-top:4px solid {cor};border-radius:4px;'
            f'padding:16px 12px 12px;text-align:center;">'
            f'<div style="font-size:10px;font-weight:700;text-transform:uppercase;'
            f'letter-spacing:.5px;color:#888;margin-bottom:8px;">{label}</div>'
            f'<div style="font-size:{fs}px;font-weight:800;{val_style}line-height:1.1;">{val}</div>'
            f'<div style="font-size:10px;color:#aaa;margin-top:4px;">{sub}</div>'
            f'</div></td>'
        )

    # Banner de aviso quando não houve casos ontem (modo fallback)
    banner_fallback = (
        '<div style="background:#FFF3CD;border:1px solid #FFEAA7;border-radius:4px;'
        'padding:12px 16px;margin-bottom:20px;font-size:13px;color:#856404;">'
        '⚠️ <strong>Sem casos registrados ontem.</strong> '
        'Segue abaixo o resumo dos últimos 7 dias para referência.</div>'
    ) if fallback else ''

    corpo_texto = (
        'Não foram registrados avisos de sinistro ontem. '
        'Para referência, segue o resumo dos últimos 7 dias. '
        'Os casos estão destacados na planilha Excel em anexo.'
    ) if fallback else (
        'Segue o resumo dos avisos de sinistro registrados ontem. '
        'Os dados completos — incluindo vigência da apólice, setor e grupo econômico — '
        'estão na planilha Excel em anexo.'
    )

    # Cards — linha 1
    if fallback:
        cards_l1 = (
            card('#E65100', 'Ontem', '0', 'nenhum caso', destaque_cor='#E65100') +
            card('#0071CE', 'Última Semana', qtd, 'casos nos 7 dias') +
            card('#00C4B4', 'Valor Total', formatar_brl(total_v), 'soma da semana')
        )
    else:
        cards_l1 = (
            card('#0071CE', 'Total de Casos', qtd, 'sinistros ontem') +
            card('#00C4B4', 'Valor Total', formatar_brl(total_v), 'soma dos sinistrados') +
            card('#7DC242', 'Maior Caso', formatar_brl(maior_v), maior_nome)
        )

    # Cards — linha 2 (só no fallback, mostra Maior Caso)
    cards_l2 = (
        card('#7DC242', 'Maior Caso', formatar_brl(maior_v), maior_nome, '50%')
    ) if fallback else ''

    return (
        '<!DOCTYPE html><html lang="pt-BR"><head><meta charset="UTF-8"></head>'
        '<body style="margin:0;padding:40px 20px;background:#f0f2f5;font-family:Arial,Helvetica,sans-serif;">'
        '<div style="max-width:620px;margin:0 auto;background:#fff;border-radius:4px;overflow:hidden;'
        'box-shadow:0 2px 12px rgba(0,0,0,.12);">'
        f'<div style="background:#0071CE;padding:24px 40px;text-align:center;">{img_h}</div>'
        '<div style="height:6px;background:linear-gradient(to right,#0071CE 0%,#0071CE 30%,'
        '#00A3D9 30%,#00A3D9 55%,#00C4B4 55%,#00C4B4 75%,#7DC242 75%,#7DC242 100%);"></div>'
        '<div style="padding:36px 40px 28px;">'
        '<p style="font-size:14px;color:#444;margin:0 0 8px;">Olá, equipe de Crédito,</p>'
        '<h1 style="font-size:22px;font-weight:700;color:#0071CE;margin:0 0 6px;">'
        'Relatório Diário de Sinistros</h1>'
        f'<p style="font-size:13px;color:#777;margin:0 0 24px;">Período: {label_periodo}</p>'
        f'{banner_fallback}'
        f'<p style="font-size:14px;color:#444;line-height:1.6;margin:0 0 28px;">{corpo_texto}</p>'
        f'<table width="100%" cellpadding="0" cellspacing="0" style="margin-bottom:{"12px" if cards_l2 else "28px"};"><tr>{cards_l1}</tr></table>'
        + (f'<table width="100%" cellpadding="0" cellspacing="0" style="margin-bottom:28px;"><tr>{cards_l2}</tr></table>' if cards_l2 else '')
        + '<div style="background:#f5f8ff;border:1px solid #d0dff5;border-radius:4px;'
        'padding:14px 18px;margin-bottom:28px;">'
        + f'<strong style="color:#1D6F42;display:block;font-size:13px;margin-bottom:4px;">{nome_arquivo}</strong>'
        + (
            f'<span style="font-size:13px;color:#444;">'
            f'Planilha com <strong>{total_ano}</strong> casos em 2026 · '
            f'destacados: {qtd} {"ontem" if not fallback else "nos últimos 7 dias"} · '
            f'Nº Sinistro, Data, Segurado, Devedor, CNPJ, Valor, Setor e Subsetor</span>'
            if total_ano and total_ano > qtd else
            f'<span style="font-size:13px;color:#444;">Planilha com {qtd} casos · '
            f'Nº Sinistro, Data, Segurado, Devedor, CNPJ, Valor, Setor e Subsetor</span>'
        )
        + '</div>'
        '<hr style="border:none;border-top:1px solid #eee;margin-bottom:20px;">'
        '<p style="font-size:13px;color:#666;line-height:1.7;margin:0;">'
        'Este relatório é gerado automaticamente todo dia às 09h (BRT).<br>'
        'Os dados também são registrados na base online Google Sheets.</p>'
        '</div>'
        f'<div style="background:#0071CE;padding:20px 40px;text-align:center;">{img_f}'
        '<div style="font-size:11px;color:rgba(255,255,255,.6);margin-top:10px;">'
        'Mensagem automática — não é necessário responder</div>'
        '</div>'
        '</div></body></html>'
    )


def enviar_relatorio(buf: BytesIO, qtd: int, registros: list, ontem: date,
                     fallback: bool = False, total_ano: int = 0):
    label_periodo = ontem.strftime('%d/%m/%Y')
    nome_arquivo  = f"Sinistros_Diario_{ontem.strftime('%d%m%Y')}.xlsx"

    logo     = _logo_path()
    tem_logo = os.path.exists(logo)

    if fallback:
        plain_text = (
            f"Olá equipe,\n\nSem sinistros registrados ontem ({label_periodo}).\n"
            f"Segue o resumo dos últimos 7 dias — {qtd} casos.\n\nGerado automaticamente às 09h BRT.\n"
        )
        assunto = f'Relatório Diário de Sinistros — {label_periodo} (0 ontem | {qtd} na semana)'
    else:
        plain_text = (
            f"Olá equipe,\n\nSegue o relatório diário de sinistros de {label_periodo}.\n"
            f"Total: {qtd} sinistros.\n\nGerado automaticamente às 09h BRT.\n"
        )
        assunto = f'Relatório Diário de Sinistros — {label_periodo} ({qtd} casos)'

    msg = MIMEMultipart('mixed')
    msg['From']    = EMAIL_CAIXA
    msg['To']      = ', '.join(DESTINATARIOS)
    msg['Subject'] = assunto

    related = MIMEMultipart('related')
    related.attach(MIMEText(
        gerar_html_email(qtd, registros, label_periodo, nome_arquivo,
                         tem_logo, fallback=fallback, total_ano=total_ano),
        'html', 'utf-8'
    ))
    if tem_logo:
        img = MIMEImage(open(logo, 'rb').read(), 'png')
        img.add_header('Content-ID', '<avla_logo>')
        img.add_header('Content-Disposition', 'inline', filename='avla_logo.png')
        related.attach(img)

    alt = MIMEMultipart('alternative')
    alt.attach(MIMEText(plain_text, 'plain', 'utf-8'))
    alt.attach(related)
    msg.attach(alt)

    parte = MIMEBase('application', 'octet-stream')
    parte.set_payload(buf.read())
    encoders.encode_base64(parte)
    parte.add_header('Content-Disposition', f'attachment; filename="{nome_arquivo}"')
    msg.attach(parte)

    with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as servidor:
        servidor.starttls()
        servidor.login(EMAIL_CAIXA, APP_PASSWORD)
        servidor.sendmail(EMAIL_CAIXA, DESTINATARIOS, msg.as_string())

    print(f"Email enviado para {', '.join(DESTINATARIOS)} — {qtd} sinistros.")


# ─────────────────────────────────────────────
#  Main
# ─────────────────────────────────────────────

def main():
    agora = datetime.now()
    print(f"[{agora.strftime('%d/%m/%Y %H:%M')}] Iniciando relatório diário AVLA...")

    ontem     = date.today() - timedelta(days=1)
    ontem_str = ontem.strftime('%d/%m/%Y')

    # Coleta SEMPRE todos os registros de 2026 até hoje
    inicio_ano = datetime(2026, 1, 1, 0, 0, 0)
    fim_hoje   = datetime(date.today().year, date.today().month, date.today().day, 23, 59, 59)

    print(f"Buscando 2026 completo até {date.today().strftime('%d/%m/%Y')}...")
    registros_ano = coletar_emails(inicio_ano, fim_hoje)

    if not registros_ano:
        print("Nenhum sinistro encontrado em 2026. Email não enviado.")
        return

    # Determina modo de exibição: D-1 normal ou fallback (7 dias)
    registros_ontem = [r for r in registros_ano if r.get('Data') == ontem_str]
    fallback = len(registros_ontem) == 0

    if fallback:
        print(f"Nenhum sinistro em {ontem_str} — modo fallback (últimos 7 dias).")
    else:
        print(f"{len(registros_ontem)} sinistro(s) em {ontem_str} — modo normal.")

    # Enriquece TODOS os registros do ano
    lookup_cnae, lookup_grupo = carregar_tabelas_auxiliares()
    registros_ano = enriquecer(registros_ano, lookup_cnae, lookup_grupo)

    # Escreve na base online — também lê e mescla casos manuais da 2ª aba
    escrever_sheets(registros_ano, lookup_cnae, lookup_grupo)

    # Re-filtra registros para exibição nos cards (após enriquecimento)
    # Nota: cards/email usam apenas os casos de email (não os manuais)
    if not fallback:
        registros_display = [r for r in registros_ano if r.get('Data') == ontem_str]
    else:
        corte_7d = date.today() - timedelta(days=7)
        registros_display = [r for r in registros_ano
                             if _parse_date(r.get('Data', '')) >= corte_7d]

    # Excel com TODOS os registros (email + manuais), ordenados por data
    # Para isso, lemos os manuais de novo (já enriquecidos, sem conexão Sheets)
    # Usamos o registros_ano já enriquecido; manuais serão incluídos pelo Sheets mas
    # o Excel só tem email para não duplicar (já que Sheets tem a visão completa)
    excel = gerar_excel(registros_ano, ontem, fallback=fallback)

    # Email: cards do período de display; descrição do Excel mostra total do ano
    enviar_relatorio(excel, len(registros_display), registros_display, ontem,
                     fallback=fallback, total_ano=len(registros_ano))

    print("Concluído com sucesso.")


if __name__ == '__main__':
    main()
