#!/usr/bin/env python3
"""
Teste Histórico de Sinistros - AVLA Crédito
Coleta TODOS os avisos de sinistro de 01/01/2026 até hoje.
Uso exclusivo para validação — envia apenas para lsilva@avla.com.
"""

import imaplib
import email
import smtplib
import re
import os
import sys
import pandas as pd
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email.mime.image import MIMEImage
from email import encoders
from datetime import datetime, timedelta
from bs4 import BeautifulSoup
from io import BytesIO
from openpyxl.styles import Font, PatternFill, Alignment

EMAIL_CAIXA       = os.environ.get("EMAIL_CAIXA",       "mgignon@avla.com")
APP_PASSWORD      = os.environ.get("APP_PASSWORD",       "")
ASSUNTO_FILTRO    = os.environ.get("ASSUNTO_FILTRO",    "SINISTRO")
REMETENTE_FILTRO  = os.environ.get("REMETENTE_FILTRO",  "notificaciones-01@avla.com")
EMAIL_DESTINO     = "lsilva@avla.com"

INICIO = datetime(2026, 1, 1)
FIM    = datetime.now().replace(hour=23, minute=59, second=59)

IMAP_SERVER = "imap.gmail.com"
SMTP_SERVER = "smtp.gmail.com"
SMTP_PORT   = 587

if not APP_PASSWORD:
    print("ERRO: variável de ambiente APP_PASSWORD não definida.")
    sys.exit(1)


def _select(mail, nome):
    """Tenta selecionar uma pasta. Retorna (True, contagem) se OK."""
    try:
        nome_imap = f'"{nome}"' if any(c in nome for c in '[] ') else nome
        typ, data = mail.select(nome_imap)
        if typ == 'OK':
            count = data[0].decode() if data and data[0] else '?'
            return True, count
        return False, 0
    except Exception:
        return False, 0


def selecionar_pasta_completa(mail):
    """Seleciona All Mail do Gmail (busca mais abrangente)."""
    # Log de todas as pastas disponíveis para diagnóstico
    try:
        _, pastas = mail.list()
        print("=== Pastas disponíveis ===")
        for item in pastas:
            if item:
                print(" ", item.decode('utf-8') if isinstance(item, bytes) else item)
        print("=========================")
    except Exception as e:
        print(f"Erro ao listar pastas: {e}")

    # 1) Detecta via atributo \All (nome varia por idioma da conta)
    try:
        _, pastas = mail.list()
        for item in pastas:
            if item is None:
                continue
            decoded = item.decode('utf-8') if isinstance(item, bytes) else str(item)
            if '\\All' in decoded:
                # Extrai nome no final da linha
                m = re.search(r'"([^"]+)"\s*$|(\S+)\s*$', decoded)
                if m:
                    nome = (m.group(1) or m.group(2)).strip()
                    ok, count = _select(mail, nome)
                    if ok:
                        print(f"✓ Pasta selecionada (All): {nome} — {count} msgs")
                        return
    except Exception as e:
        print(f"Erro ao detectar All Mail: {e}")

    # 2) Fallback — All Mail explícito antes de INBOX
    for nome in ['[Gmail]/All Mail', '[Gmail]/Todos os e-mails', 'INBOX']:
        ok, count = _select(mail, nome)
        if ok:
            print(f"✓ Pasta selecionada (fallback): {nome} — {count} msgs")
            return

    raise RuntimeError("Não foi possível selecionar nenhuma pasta de email.")


def _detectar_lixeira(mail):
    """Detecta o nome da pasta Lixeira/Trash do Gmail (varia por idioma da conta)."""
    try:
        _, pastas = mail.list()
        for item in pastas:
            if item is None:
                continue
            decoded = item.decode('utf-8') if isinstance(item, bytes) else str(item)
            if '\\Trash' in decoded:
                m = re.search(r'"([^"]+)"\s*$|(\S+)\s*$', decoded)
                if m:
                    return (m.group(1) or m.group(2)).strip()
    except Exception as e:
        print(f"Erro ao detectar Lixeira: {e}")
    return None


def _buscar_pasta(mail, since, before, vistos):
    """Busca sinistros na pasta atualmente selecionada.
    Retorna (lista_registros, set_vistos_atualizado) — deduplicação por Message-ID.
    """
    try:
        _, ids = mail.search(None, f'(SUBJECT "{ASSUNTO_FILTRO}" SINCE "{since}" BEFORE "{before}")')
    except Exception as e:
        print(f"Erro ao buscar na pasta: {e}")
        return [], vistos

    novos = []
    for msg_id in ids[0].split():
        try:
            _, dados = mail.fetch(msg_id, '(RFC822)')
            msg = email.message_from_bytes(dados[0][1])
        except Exception:
            continue

        mid = msg.get('Message-ID', f'_noid_{msg_id.decode()}_')
        if mid in vistos:
            continue
        vistos.add(mid)

        remetente = str(msg.get('From', ''))
        if REMETENTE_FILTRO and REMETENTE_FILTRO.lower() not in remetente.lower():
            continue

        assunto    = str(msg.get('Subject', ''))
        data_email = email.utils.parsedate_to_datetime(msg['Date'])

        html_body = None
        for part in msg.walk():
            if part.get_content_type() == 'text/html':
                payload   = part.get_payload(decode=True)
                charset   = part.get_content_charset() or 'utf-8'
                html_body = payload.decode(charset, errors='ignore')
                break

        if not html_body:
            continue

        campos      = parse_corpo_email(html_body)
        num_assunto = extrair_numero_sinistro(assunto)
        num_corpo   = campos.get('N° Sinistro', '') or _extrair_num_corpo(html_body)
        num_final   = num_assunto or num_corpo
        if not num_final:
            print(f"  ⚠ N° Sinistro não encontrado | assunto: {assunto[:80]}")
        campos['N° Sinistro'] = num_final
        campos['Data']        = data_email.strftime('%d/%m/%Y')
        novos.append(campos)

    return novos, vistos


def extrair_numero_sinistro(assunto: str) -> str:
    """Tenta extrair o número do sinistro do ASSUNTO do email."""
    padroes = [
        r'N[°º]?\s*(?:DE\s+)?S[Ii][Nn][Ii][Ee]?[Ss][Tt][Rr][Oo]\s*[:\-]?\s*(\d+)',
        r'S[Ii][Nn][Ii][Ee]?[Ss][Tt][Rr][Oo]\s*[N°nº#]?\s*[:\-]?\s*(\d+)',
        r'\b(\d{6,})\b',
    ]
    for p in padroes:
        m = re.search(p, assunto, re.IGNORECASE)
        if m:
            return m.group(1)
    return ''


def _extrair_num_corpo(html: str) -> str:
    """Extrai número do sinistro direto do corpo do email via regex.
    Usado como fallback quando o assunto e os campos bold não encontram o número.
    """
    soup = BeautifulSoup(html, 'html.parser')
    texto = soup.get_text(separator=' ')

    padroes = [
        r'N[°º\.]\s*(?:de\s+)?[Ss]ini[es]stro\s*[:\-]?\s*(\d{4,})',
        r'[Ss]ini[es]stro\s*[N°nº#\.]*\s*[:\-]?\s*(\d{4,})',
        r'(?:N[°º]|No\.?|Nro\.?|Nr\.?)\s*[:\-]?\s*(\d{5,})',
        r'\b(\d{8,})\b',
    ]
    for p in padroes:
        m = re.search(p, texto, re.IGNORECASE)
        if m:
            return m.group(1)
    return ''


def parse_corpo_email(html: str) -> dict:
    soup = BeautifulSoup(html, 'html.parser')
    texto = soup.get_text(separator='\n')

    campos_aliases = {
        'N° Sinistro':      ['N° Sinistro', 'Nº Sinistro', 'N° de Sinistro',
                             'Número do Sinistro', 'Número de Sinistro',
                             'N° de Siniestro', 'Nº de Siniestro',
                             'Número de Siniestro', 'No. de Siniestro',
                             'No de Siniestro', 'Siniestro'],
        'Segurado':         ['Segurado'],
        'Filial':           ['Filial'],
        'Apólice':          ['Apólice', 'Apolice', 'Póliza', 'Poliza'],
        'Devedor':          ['Devedor', 'Deudor'],
        'CNPJ do Devedor':  ['CNPJ do devedor', 'CNPJ do Devedor', 'CNPJ',
                             'RUT', 'RUT del Deudor'],
        'Ocorrência':       ['Ocorrência', 'Ocorrencia', 'Siniestro',
                             'Tipo de Siniestro'],
        'Declaração':       ['Declaração', 'Declaracao', 'Fecha de Declaración',
                             'Fecha Declaracion'],
        'Valor Sinistrado': ['Valor sinistrado', 'Valor Sinistrado',
                             'Monto Sinistrado', 'Monto del Siniestro',
                             'Valor do Sinistro'],
    }

    resultado = {}

    for tag in soup.find_all(['b', 'strong']):
        label = tag.get_text(strip=True).rstrip(':')
        for campo, aliases in campos_aliases.items():
            if campo in resultado:
                continue
            for alias in aliases:
                if alias.lower() == label.lower():
                    proximo = tag.next_sibling
                    if proximo:
                        if hasattr(proximo, 'get_text'):
                            valor = proximo.get_text(strip=True).lstrip(':').strip()
                        else:
                            valor = str(proximo).strip().lstrip(':').strip()
                        if valor:
                            resultado[campo] = valor
                    break

    for campo, aliases in campos_aliases.items():
        if campo in resultado:
            continue
        for alias in aliases:
            padrao = re.compile(rf'{re.escape(alias)}\s*:?\s*(.+)', re.IGNORECASE)
            m = padrao.search(texto)
            if m:
                resultado[campo] = m.group(1).strip()
                break

    return resultado


def coletar_emails() -> list:
    print(f"Conectando em {EMAIL_CAIXA}...")
    print(f"Período: {INICIO.strftime('%d/%m/%Y')} a {FIM.strftime('%d/%m/%Y')}")

    mail = imaplib.IMAP4_SSL(IMAP_SERVER)
    mail.login(EMAIL_CAIXA, APP_PASSWORD)

    since  = INICIO.strftime("%d-%b-%Y")
    before = (FIM + timedelta(days=1)).strftime("%d-%b-%Y")

    vistos    = set()
    registros = []

    # 1) All Mail (emails ativos + arquivados)
    selecionar_pasta_completa(mail)
    novos, vistos = _buscar_pasta(mail, since, before, vistos)
    print(f"All Mail: {len(novos)} sinistros encontrados")
    registros.extend(novos)

    # 2) Lixeira (emails descartados acidentalmente)
    nome_lixeira = _detectar_lixeira(mail)
    if nome_lixeira:
        ok, count = _select(mail, nome_lixeira)
        if ok:
            print(f"Lixeira ({nome_lixeira}): {count} msgs — buscando sinistros...")
            novos, vistos = _buscar_pasta(mail, since, before, vistos)
            print(f"Lixeira: +{len(novos)} sinistros adicionais")
            registros.extend(novos)
    else:
        print("Pasta Lixeira não detectada — pulando.")

    mail.logout()
    print(f"Total coletado: {len(registros)} sinistros")
    return registros


def gerar_excel(registros: list) -> BytesIO:
    import openpyxl

    colunas = [
        'ID', 'N° Sinistro', 'Data', 'Segurado', 'Filial', 'Apólice',
        'Devedor', 'CNPJ do Devedor', 'Ocorrência', 'Declaração', 'Valor Sinistrado'
    ]

    df = pd.DataFrame(registros)
    df.insert(0, 'ID', range(1, len(df) + 1))

    for col in colunas:
        if col not in df.columns:
            df[col] = ''

    df = df[colunas]

    buf = BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Sinistros')
        ws = writer.sheets['Sinistros']

        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(fill_type='solid', fgColor='003087')
        for cell in ws[1]:
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

        ws.row_dimensions[1].height = 20

        for col in ws.columns:
            max_len = max((len(str(c.value or '')) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 55)

        # Destaque: linhas da última semana em azul claro
        hoje_d       = datetime.now().date()
        semana_atras = hoje_d - timedelta(days=7)
        destaque     = PatternFill(fill_type='solid', fgColor='DCF0FF')

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            try:
                data_str = row[2].value  # coluna 'Data'
                if data_str and datetime.strptime(str(data_str), '%d/%m/%Y').date() >= semana_atras:
                    for cell in row:
                        cell.fill = destaque
            except (ValueError, TypeError):
                pass

    buf.seek(0)
    return buf


def _logo_path():
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), 'avla_logo.png')


def parse_valor(s: str) -> float:
    try:
        s = re.sub(r'[^\d,]', '', str(s))
        return float(s.replace(',', '.')) if s else 0.0
    except Exception:
        return 0.0


def formatar_brl(v: float) -> str:
    if v >= 1_000_000:
        return 'R$ ' + '{:.1f}'.format(v / 1_000_000).replace('.', ',') + 'M'
    if v >= 1_000:
        return 'R$ ' + '{:.0f}'.format(v / 1_000) + 'K'
    return 'R$ ' + '{:.2f}'.format(v).replace('.', ',')


def gerar_html_email(qtd, registros, label_periodo, nome_arquivo, tem_logo=True):
    img_h = ('<img src="cid:avla_logo" alt="AVLA" style="height:52px;display:block;margin:0 auto;">'
             if tem_logo else '<span style="color:white;font-size:26px;font-weight:bold;">AVLA</span>')
    img_f = ('<img src="cid:avla_logo" alt="AVLA" style="height:36px;display:block;margin:0 auto;">'
             if tem_logo else '<span style="color:white;font-size:18px;font-weight:bold;">AVLA</span>')

    total_v = sum(parse_valor(r.get('Valor Sinistrado', '')) for r in registros)
    maior_v = max((parse_valor(r.get('Valor Sinistrado', '')) for r in registros), default=0.0)
    maior_r = next((r for r in registros if parse_valor(r.get('Valor Sinistrado', '')) == maior_v), {})
    maior_nome = maior_r.get('Devedor', maior_r.get('Segurado', '—'))
    if len(maior_nome) > 22:
        maior_nome = maior_nome[:22] + '…'

    hoje         = datetime.now()
    semana_atras = hoje - timedelta(days=7)
    mes_atras    = hoje - timedelta(days=30)

    def _parse_dt(s):
        try:
            return datetime.strptime(s, '%d/%m/%Y')
        except Exception:
            return None

    datas        = [_parse_dt(r.get('Data', '')) for r in registros]
    casos_semana = sum(1 for d in datas if d and d >= semana_atras)
    casos_mes    = sum(1 for d in datas if d and d >= mes_atras)

    def card(cor, label, val, sub, width='33%'):
        fs = '22' if label == 'Total de Casos' else '18'
        return (
            f'<td width="{width}" style="padding:0 5px;">'
            f'<div style="border:1px solid #e0e0e0;border-top:4px solid {cor};border-radius:4px;padding:16px 12px 12px;text-align:center;">'
            f'<div style="font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:.5px;color:#888;margin-bottom:8px;">{label}</div>'
            f'<div style="font-size:{fs}px;font-weight:800;color:#1a1a1a;line-height:1.1;">{val}</div>'
            f'<div style="font-size:10px;color:#aaa;margin-top:4px;">{sub}</div>'
            f'</div></td>'
        )

    return (
        '<!DOCTYPE html><html lang="pt-BR"><head><meta charset="UTF-8"></head>'
        '<body style="margin:0;padding:40px 20px;background:#f0f2f5;font-family:Arial,Helvetica,sans-serif;">'
        '<div style="max-width:620px;margin:0 auto;background:#fff;border-radius:4px;overflow:hidden;box-shadow:0 2px 12px rgba(0,0,0,.12);">'
        f'<div style="background:#0071CE;padding:24px 40px;text-align:center;">{img_h}</div>'
        '<div style="height:6px;background:linear-gradient(to right,#0071CE 0%,#0071CE 30%,#00A3D9 30%,#00A3D9 55%,#00C4B4 55%,#00C4B4 75%,#7DC242 75%,#7DC242 100%);"></div>'
        '<div style="padding:36px 40px 28px;">'
        '<p style="font-size:14px;color:#444;margin:0 0 8px;">Olá, equipe de Crédito,</p>'
        '<h1 style="font-size:22px;font-weight:700;color:#0071CE;margin:0 0 6px;">Histórico de Sinistros 2026</h1>'
        f'<p style="font-size:13px;color:#777;margin:0 0 24px;">Período: {label_periodo}</p>'
        '<p style="font-size:14px;color:#444;line-height:1.6;margin:0 0 28px;">'
        'Segue o resumo dos avisos de sinistro registrados no período. Os dados completos estão na planilha Excel em anexo.</p>'
        '<table width="100%" cellpadding="0" cellspacing="0" style="margin-bottom:12px;"><tr>'
        f'{card("#0071CE","Total de Casos",qtd,"sinistros encontrados")}'
        f'{card("#00C4B4","Valor Total",formatar_brl(total_v),"soma dos sinistrados")}'
        f'{card("#7DC242","Maior Caso",formatar_brl(maior_v),maior_nome)}'
        '</tr></table>'
        '<table width="100%" cellpadding="0" cellspacing="0" style="margin-bottom:28px;"><tr>'
        f'{card("#00A3D9","Última Semana",casos_semana,"últimos 7 dias","50%")}'
        f'{card("#003087","Último Mês",casos_mes,"últimos 30 dias","50%")}'
        '</tr></table>'
        '<div style="background:#f5f8ff;border:1px solid #d0dff5;border-radius:4px;padding:14px 18px;margin-bottom:28px;">'
        f'<strong style="color:#1D6F42;display:block;font-size:13px;margin-bottom:4px;">{nome_arquivo}</strong>'
        f'<span style="font-size:13px;color:#444;">Planilha com todos os {qtd} casos · '
        'ID, Nº Sinistro, Data, Segurado, Devedor, CNPJ, Valor</span>'
        '</div>'
        '<hr style="border:none;border-top:1px solid #eee;margin-bottom:20px;">'
        '<p style="font-size:13px;color:#666;line-height:1.7;margin:0;">Este é um relatório de validação gerado sob demanda.</p>'
        '</div>'
        f'<div style="background:#0071CE;padding:20px 40px;text-align:center;">{img_f}'
        '<div style="font-size:11px;color:rgba(255,255,255,.6);margin-top:10px;">Mensagem automática — não é necessário responder</div>'
        '</div>'
        '</div></body></html>'
    )


def enviar_relatorio(buf: BytesIO, qtd: int, registros: list):
    label_periodo = f"{INICIO.strftime('%d/%m/%Y')} a {FIM.strftime('%d/%m/%Y')}"
    nome_arquivo  = f"Sinistros_Historico_2026_{FIM.strftime('%d%m%Y')}.xlsx"

    logo     = _logo_path()
    tem_logo = os.path.exists(logo)

    plain_text = (
        f"Lucas,\n\nEmail de validação do coletor de sinistros.\n"
        f"Período: {label_periodo}\nTotal: {qtd} sinistros.\n\n— Automação AVLA Crédito\n"
    )

    msg = MIMEMultipart('mixed')
    msg['From']    = EMAIL_CAIXA
    msg['To']      = EMAIL_DESTINO
    msg['Subject'] = f'[TESTE] Histórico de Sinistros 2026 — {qtd} casos encontrados'

    related = MIMEMultipart('related')
    related.attach(MIMEText(
        gerar_html_email(qtd, registros, label_periodo, nome_arquivo, tem_logo),
        'html', 'utf-8'
    ))
    if tem_logo:
        img = MIMEImage(open(logo, 'rb').read(), 'png')
        img.add_header('Content-ID', '<avla_logo>')
        img.add_header('Content-Disposition', 'inline', filename='avla_logo.png')
        related.attach(img)

    alt = MIMEMultipart('alternative')
    alt.attach(MIMEText(plain_text, 'plain', 'utf-8'))
    alt.attach(related)
    msg.attach(alt)

    parte = MIMEBase('application', 'octet-stream')
    parte.set_payload(buf.read())
    encoders.encode_base64(parte)
    parte.add_header('Content-Disposition', f'attachment; filename="{nome_arquivo}"')
    msg.attach(parte)

    with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as servidor:
        servidor.starttls()
        servidor.login(EMAIL_CAIXA, APP_PASSWORD)
        servidor.send_message(msg)

    print(f"Email enviado para {EMAIL_DESTINO} — {qtd} sinistros.")


def main():
    print(f"[{datetime.now().strftime('%d/%m/%Y %H:%M')}] Iniciando coleta histórica 2026...")

    registros = coletar_emails()

    if not registros:
        print("Nenhum sinistro encontrado no período.")
        return

    excel = gerar_excel(registros)
    enviar_relatorio(excel, len(registros), registros)
    print("Concluído.")


if __name__ == '__main__':
    main()
