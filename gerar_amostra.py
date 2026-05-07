import pandas as pd
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

dados = [
    {
        "N° Sinistro":     "552026000807",
        "Data":            "06/02/2026",
        "Segurado":        "Química Amparo Ltda",
        "Filial":          "Química Amparo Ltda",
        "Apólice":         "1002026000809 – Seguro de Crédito",
        "Devedor":         "Adriana Distribuidora de Produtos Alimentícios Ltda",
        "CNPJ do Devedor": "04.724.476/0001-28",
        "Ocorrência":      "31/10/2025",
        "Declaração":      "06/02/2026",
        "Valor Sinistrado":"R$ 1.245.942,09",
    },
    {
        "N° Sinistro":     "552026001134",
        "Data":            "07/02/2026",
        "Segurado":        "Frigorífico Bela Vista S.A.",
        "Filial":          "Frigorífico Bela Vista S.A.",
        "Apólice":         "1002026000412 – Seguro de Crédito",
        "Devedor":         "Mercado Central Atacadista Ltda",
        "CNPJ do Devedor": "12.345.678/0001-90",
        "Ocorrência":      "15/11/2025",
        "Declaração":      "07/02/2026",
        "Valor Sinistrado":"R$ 387.500,00",
    },
    {
        "N° Sinistro":     "552026001290",
        "Data":            "10/02/2026",
        "Segurado":        "Têxtil Norte S.A.",
        "Filial":          "Têxtil Norte S.A. – Filial SP",
        "Apólice":         "1002025008873 – Seguro de Crédito",
        "Devedor":         "Confecções Irmãos Pereira Ltda",
        "CNPJ do Devedor": "98.765.432/0001-11",
        "Ocorrência":      "20/10/2025",
        "Declaração":      "10/02/2026",
        "Valor Sinistrado":"R$ 210.000,00",
    },
    {
        "N° Sinistro":     "552026001455",
        "Data":            "11/02/2026",
        "Segurado":        "Agro Cerrado Exportações Ltda",
        "Filial":          "Agro Cerrado Exportações Ltda",
        "Apólice":         "1002026000190 – Seguro de Crédito",
        "Devedor":         "Grãos do Sul Comércio e Logística S.A.",
        "CNPJ do Devedor": "55.123.987/0001-44",
        "Ocorrência":      "05/12/2025",
        "Declaração":      "11/02/2026",
        "Valor Sinistrado":"R$ 892.300,00",
    },
    {
        "N° Sinistro":     "552026001601",
        "Data":            "12/02/2026",
        "Segurado":        "Indústria Metálica Fortaleza Ltda",
        "Filial":          "Indústria Metálica Fortaleza Ltda",
        "Apólice":         "1002025007741 – Seguro de Crédito",
        "Devedor":         "Construções & Reformas Rápidas Eireli",
        "CNPJ do Devedor": "33.456.789/0001-05",
        "Ocorrência":      "01/11/2025",
        "Declaração":      "12/02/2026",
        "Valor Sinistrado":"R$ 134.800,00",
    },
    {
        "N° Sinistro":     "552026001788",
        "Data":            "13/02/2026",
        "Segurado":        "Distribuidora Paulista de Alimentos S.A.",
        "Filial":          "Distribuidora Paulista de Alimentos S.A.",
        "Apólice":         "1002026000567 – Seguro de Crédito",
        "Devedor":         "Supermercados Família Unida Ltda",
        "CNPJ do Devedor": "77.890.123/0001-67",
        "Ocorrência":      "18/11/2025",
        "Declaração":      "13/02/2026",
        "Valor Sinistrado":"R$ 508.750,00",
    },
    {
        "N° Sinistro":     "552026001934",
        "Data":            "14/02/2026",
        "Segurado":        "Laticínios Serra Gaúcha S.A.",
        "Filial":          "Laticínios Serra Gaúcha S.A. – Filial RS",
        "Apólice":         "1002025006230 – Seguro de Crédito",
        "Devedor":         "Atacadão do Leite Comércio Ltda",
        "CNPJ do Devedor": "44.321.654/0001-23",
        "Ocorrência":      "10/12/2025",
        "Declaração":      "14/02/2026",
        "Valor Sinistrado":"R$ 72.400,00",
    },
    {
        "N° Sinistro":     "552026002105",
        "Data":            "17/02/2026",
        "Segurado":        "Importadora Global Tech Ltda",
        "Filial":          "Importadora Global Tech Ltda",
        "Apólice":         "1002026000834 – Seguro de Crédito",
        "Devedor":         "Revendas Conectadas do Brasil S.A.",
        "CNPJ do Devedor": "66.789.012/0001-89",
        "Ocorrência":      "22/10/2025",
        "Declaração":      "17/02/2026",
        "Valor Sinistrado":"R$ 1.780.600,00",
    },
]

colunas = [
    "ID", "N° Sinistro", "Data", "Segurado", "Filial", "Apólice",
    "Devedor", "CNPJ do Devedor", "Ocorrência", "Declaração", "Valor Sinistrado"
]

df = pd.DataFrame(dados)
df.insert(0, "ID", range(1, len(df) + 1))
df = df[colunas]

caminho = "Sinistros_0902_15022026_AMOSTRA.xlsx"

with pd.ExcelWriter(caminho, engine="openpyxl") as writer:
    df.to_excel(writer, index=False, sheet_name="Sinistros")
    ws = writer.sheets["Sinistros"]

    # Cabeçalho
    header_font = Font(bold=True, color="FFFFFF", size=11, name="Arial")
    header_fill = PatternFill(fill_type="solid", fgColor="003087")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="FFFFFF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        cell.border    = border

    ws.row_dimensions[1].height = 28

    # Linhas de dados — zebra
    fill_par  = PatternFill(fill_type="solid", fgColor="EBF2FA")
    fill_impar = PatternFill(fill_type="solid", fgColor="FFFFFF")
    data_font  = Font(name="Arial", size=10)
    data_align_center = Alignment(horizontal="center", vertical="center")
    data_align_left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        fill = fill_par if row_idx % 2 == 0 else fill_impar
        for cell in row:
            cell.fill      = fill
            cell.font      = data_font
            col_letter = cell.column_letter
            # Colunas de texto longo: alinhamento esquerdo
            if cell.column in [4, 5, 6, 7]:  # Segurado, Filial, Apólice, Devedor
                cell.alignment = data_align_left
            else:
                cell.alignment = data_align_center
        ws.row_dimensions[row_idx].height = 22

    # Larguras das colunas
    larguras = {
        "A": 6,   # ID
        "B": 16,  # N° Sinistro
        "C": 12,  # Data
        "D": 28,  # Segurado
        "E": 28,  # Filial
        "F": 36,  # Apólice
        "G": 38,  # Devedor
        "H": 20,  # CNPJ
        "I": 12,  # Ocorrência
        "J": 12,  # Declaração
        "K": 18,  # Valor Sinistrado
    }
    for col, width in larguras.items():
        ws.column_dimensions[col].width = width

    # Congelar painel no cabeçalho
    ws.freeze_panes = "A2"

print(f"Planilha gerada: {caminho}")
print(f"Linhas: {len(df)} sinistros | Colunas: {len(colunas)}")
