#!/usr/bin/env python3
"""
Coletor Semanal de Sinistros - AVLA Crédito
Coleta emails de aviso de sinistro da semana anterior e envia
relatório em Excel toda sexta-feira para a caixa de crédito.
"""

import imaplib
import email
import smtplib
import re
import os
import sys
import pandas as pd
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from datetime import datetime, timedelta
from bs4 import BeautifulSoup
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# =============================================================
# CONFIGURAÇÃO — lida de variáveis de ambiente (GitHub Secrets)
# EMAIL_CAIXA  : conta de onde LÊ os emails (sua conta pessoal)
# EMAIL_DESTINO: para quem ENVIA o relatório (caixa do time)
# =============================================================
EMAIL_CAIXA      = os.environ.get("EMAIL_CAIXA",      "mgignon@avla.com")
APP_PASSWORD     = os.environ.get("APP_PASSWORD",      "")
REMETENTE_FILTRO = os.environ.get("REMETENTE_FILTRO",  "notificaciones-01@avla.com")
EMAIL_DESTINO    = os.environ.get("EMAIL_DESTINO",     "mgignon@avla.com")

IMAP_SERVER = "imap.gmail.com"
SMTP_SERVER = "smtp.gmail.com"
SMTP_PORT   = 587

if not APP_PASSWORD:
    print("ERRO: variável de ambiente APP_PASSWORD não definida.")
    sys.exit(1)
# =============================================================


def periodo_semana_anterior():
    """Retorna (domingo, sábado) da semana anterior completa."""
    hoje = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    dias_desde_domingo = (hoje.weekday() + 1) % 7
    ultimo_domingo = hoje - timedelta(days=dias_desde_domingo + 7)
    ultimo_sabado  = ultimo_domingo + timedelta(days=6)
    return ultimo_domingo, ultimo_sabado.replace(hour=23, minute=59, second=59)


def extrair_numero_sinistro(assunto: str) -> str:
    m = re.search(r'N[°º]?\s*(?:DE\s+)?SINISTRO\s+(\d+)', assunto, re.IGNORECASE)
    if m:
        return m.group(1)
    m = re.search(r'\b(\d{12,})\b', assunto)
    return m.group(1) if m else ''


def parse_corpo_email(html: str) -> dict:
    soup = BeautifulSoup(html, 'html.parser')
    texto = soup.get_text(separator='\n')

    campos_aliases = {
        'Segurado':         ['Segurado'],
        'Filial':           ['Filial'],
        'Apólice':          ['Apólice', 'Apolice'],
        'Devedor':          ['Devedor'],
        'CNPJ do Devedor':  ['CNPJ do devedor', 'CNPJ do Devedor', 'CNPJ'],
        'Ocorrência':       ['Ocorrência', 'Ocorrencia'],
        'Declaração':       ['Declaração', 'Declaracao'],
        'Valor Sinistrado': ['Valor sinistrado', 'Valor Sinistrado'],
    }

    resultado = {}

    for tag in soup.find_all(['b', 'strong']):
        label = tag.get_text(strip=True).rstrip(':')
        for campo, aliases in campos_aliases.items():
            if campo in resultado:
                continue
            for alias in aliases:
                if alias.lower() == label.lower():
                    proximo = tag.next_sibling
                    if proximo:
                        valor = str(proximo).strip().lstrip(':').strip()
                        if valor:
                            resultado[campo] = valor
                    break

    for campo, aliases in campos_aliases.items():
        if campo in resultado:
            continue
        for alias in aliases:
            padrao = re.compile(rf'{re.escape(alias)}\s*:?\s*(.+)', re.IGNORECASE)
            m = padrao.search(texto)
            if m:
                resultado[campo] = m.group(1).strip()
                break

    return resultado


def coletar_emails(inicio: datetime, fim: datetime) -> list:
    print(f"Conectando em {EMAIL_CAIXA}...")
    mail = imaplib.IMAP4_SSL(IMAP_SERVER)
    mail.login(EMAIL_CAIXA, APP_PASSWORD)
    mail.select('inbox')

    since  = inicio.strftime("%d-%b-%Y")
    before = (fim + timedelta(days=1)).strftime("%d-%b-%Y")

    _, ids = mail.search(
        None,
        f'(FROM "{REMETENTE_FILTRO}" SINCE "{since}" BEFORE "{before}")'
    )

    ids_lista = ids[0].split()
    print(f"Emails encontrados no período: {len(ids_lista)}")

    registros = []
    for msg_id in ids_lista:
        _, dados = mail.fetch(msg_id, '(RFC822)')
        msg = email.message_from_bytes(dados[0][1])

        assunto    = str(msg.get('Subject', ''))
        data_email = email.utils.parsedate_to_datetime(msg['Date'])

        html_body = None
        for part in msg.walk():
            if part.get_content_type() == 'text/html':
                payload   = part.get_payload(decode=True)
                charset   = part.get_content_charset() or 'utf-8'
                html_body = payload.decode(charset, errors='ignore')
                break

        if not html_body:
            continue

        campos = parse_corpo_email(html_body)
        campos['N° Sinistro'] = extrair_numero_sinistro(assunto)
        campos['Data']        = data_email.strftime('%d/%m/%Y')
        registros.append(campos)

    mail.logout()
    return registros


def gerar_excel(registros: list, inicio: datetime, fim: datetime) -> BytesIO:
    colunas = [
        'ID', 'N° Sinistro', 'Data', 'Segurado', 'Filial', 'Apólice',
        'Devedor', 'CNPJ do Devedor', 'Ocorrência', 'Declaração', 'Valor Sinistrado'
    ]

    df = pd.DataFrame(registros)
    df.insert(0, 'ID', range(1, len(df) + 1))

    for col in colunas:
        if col not in df.columns:
            df[col] = ''

    df = df[colunas]

    buf = BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Sinistros')
        ws = writer.sheets['Sinistros']

        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(fill_type='solid', fgColor='003087')
        for cell in ws[1]:
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')

        ws.row_dimensions[1].height = 20

        for col in ws.columns:
            max_len = max((len(str(c.value or '')) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 55)

    buf.seek(0)
    return buf


def enviar_relatorio(buf: BytesIO, qtd: int, inicio: datetime, fim: datetime):
    msg = MIMEMultipart()
    msg['From']    = EMAIL_CAIXA
    msg['To']      = EMAIL_DESTINO
    msg['Subject'] = (
        f'Relatório Semanal de Sinistros — '
        f'{inicio.strftime("%d/%m")} a {fim.strftime("%d/%m/%Y")}'
    )

    corpo = (
        f"Olá equipe,\n\n"
        f"Segue o relatório de sinistros registrados de "
        f"{inicio.strftime('%d/%m/%Y')} a {fim.strftime('%d/%m/%Y')}.\n\n"
        f"Total de sinistros no período: {qtd}\n\n"
        f"Este relatório foi gerado automaticamente.\n"
    )
    msg.attach(MIMEText(corpo, 'plain', 'utf-8'))

    nome_arquivo = f"Sinistros_{inicio.strftime('%d%m')}_{fim.strftime('%d%m%Y')}.xlsx"
    parte = MIMEBase('application', 'octet-stream')
    parte.set_payload(buf.read())
    encoders.encode_base64(parte)
    parte.add_header('Content-Disposition', f'attachment; filename="{nome_arquivo}"')
    msg.attach(parte)

    with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as servidor:
        servidor.starttls()
        servidor.login(EMAIL_CAIXA, APP_PASSWORD)
        servidor.send_message(msg)

    print(f"Relatório enviado para {EMAIL_DESTINO} — {qtd} sinistros.")


def main():
    agora = datetime.now()
    print(f"[{agora.strftime('%d/%m/%Y %H:%M')}] Iniciando coletor de sinistros AVLA...")

    inicio, fim = periodo_semana_anterior()
    print(f"Período: {inicio.strftime('%d/%m/%Y')} (dom) a {fim.strftime('%d/%m/%Y')} (sáb)")

    registros = coletar_emails(inicio, fim)

    if not registros:
        print("Nenhum sinistro encontrado no período. Nenhum email enviado.")
        return

    excel = gerar_excel(registros, inicio, fim)
    enviar_relatorio(excel, len(registros), inicio, fim)
    print("Concluído com sucesso.")


if __name__ == '__main__':
    main()
